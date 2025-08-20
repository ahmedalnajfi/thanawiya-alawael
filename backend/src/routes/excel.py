from flask import Blueprint, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
import os
from werkzeug.utils import secure_filename

excel_bp = Blueprint('excel', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@excel_bp.route('/upload-excel', methods=['POST'])
def upload_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'لم يتم اختيار ملف'}), 400
        
        file = request.files['file']
        upload_type = request.form.get('type', 'students')
        
        if file.filename == '':
            return jsonify({'error': 'لم يتم اختيار ملف'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'نوع الملف غير مدعوم. يرجى استخدام ملفات Excel (.xlsx أو .xls)'}), 400
        
        # قراءة ملف Excel
        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active
        except Exception as e:
            return jsonify({'error': f'خطأ في قراءة ملف Excel: {str(e)}'}), 400
        
        if upload_type == 'students':
            return process_students_excel(worksheet)
        elif upload_type == 'classes':
            return process_classes_excel(worksheet)
        else:
            return jsonify({'error': 'نوع الرفع غير صحيح'}), 400
            
    except Exception as e:
        return jsonify({'error': f'خطأ في معالجة الملف: {str(e)}'}), 500

def process_students_excel(worksheet):
    """معالجة ملف Excel للطلاب"""
    required_columns = [
        'الاسم الكامل', 'رقم الهوية', 'تاريخ الميلاد', 'الصف', 'الشعبة',
        'اسم ولي الأمر', 'رقم هاتف ولي الأمر', 'العنوان', 'الجنس', 'البناية'
    ]
    
    # قراءة الصف الأول للحصول على أسماء الأعمدة
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append('')
    
    # التحقق من وجود الأعمدة المطلوبة
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        return jsonify({
            'error': f'الأعمدة التالية مفقودة: {", ".join(missing_columns)}'
        }), 400
    
    # إنشاء خريطة للأعمدة
    column_map = {}
    for i, header in enumerate(headers):
        if header in required_columns:
            column_map[header] = i
    
    students_data = []
    errors = []
    
    # قراءة البيانات من الصف الثاني فما فوق
    for row_num in range(2, worksheet.max_row + 1):
        try:
            row = worksheet[row_num]
            
            # استخراج البيانات
            name = str(row[column_map['الاسم الكامل']].value or '').strip()
            student_id = str(row[column_map['رقم الهوية']].value or '').strip()
            
            if not name or not student_id:
                errors.append(f'الصف {row_num}: الاسم أو رقم الهوية مفقود')
                continue
            
            # تحويل تاريخ الميلاد
            birth_date = None
            birth_cell = row[column_map['تاريخ الميلاد']].value
            if birth_cell:
                try:
                    if isinstance(birth_cell, str):
                        birth_date = datetime.strptime(birth_cell, '%d/%m/%Y').date()
                    elif hasattr(birth_cell, 'date'):
                        birth_date = birth_cell.date()
                    else:
                        birth_date = birth_cell
                except:
                    errors.append(f'الصف {row_num}: تنسيق تاريخ الميلاد غير صحيح')
                    continue
            
            grade = str(row[column_map['الصف']].value or '').strip()
            section = str(row[column_map['الشعبة']].value or '').strip()
            parent_name = str(row[column_map['اسم ولي الأمر']].value or '').strip()
            parent_phone = str(row[column_map['رقم هاتف ولي الأمر']].value or '').strip()
            address = str(row[column_map['العنوان']].value or '').strip()
            gender = str(row[column_map['الجنس']].value or '').strip()
            building = str(row[column_map['البناية']].value or '').strip()
            
            student_data = {
                'name': name,
                'student_id': student_id,
                'birth_date': birth_date.isoformat() if birth_date else None,
                'grade': grade,
                'section': section,
                'parent_name': parent_name,
                'parent_phone': parent_phone,
                'address': address,
                'gender': gender,
                'building': building,
                'is_french': 'فرنسي' in section.lower(),
                'created_at': datetime.now().isoformat()
            }
            
            students_data.append(student_data)
            
        except Exception as e:
            errors.append(f'الصف {row_num}: خطأ في معالجة البيانات - {str(e)}')
    
    success_count = len(students_data)
    
    response_data = {
        'success': True,
        'count': success_count,
        'type': 'students',
        'data': students_data,
        'errors': errors
    }
    
    if errors:
        response_data['warning'] = f'تم رفع {success_count} طالب بنجاح، مع {len(errors)} خطأ'
    
    return jsonify(response_data)

def process_classes_excel(worksheet):
    """معالجة ملف Excel للشعب"""
    required_columns = [
        'اسم الشعبة', 'الصف', 'النوع', 'المرحلة الدراسية', 
        'عدد الطلاب المتوقع', 'البناية', 'المعلم المسؤول'
    ]
    
    # قراءة الصف الأول للحصول على أسماء الأعمدة
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append('')
    
    # التحقق من وجود الأعمدة المطلوبة
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        return jsonify({
            'error': f'الأعمدة التالية مفقودة: {", ".join(missing_columns)}'
        }), 400
    
    # إنشاء خريطة للأعمدة
    column_map = {}
    for i, header in enumerate(headers):
        if header in required_columns:
            column_map[header] = i
    
    classes_data = []
    errors = []
    
    # قراءة البيانات من الصف الثاني فما فوق
    for row_num in range(2, worksheet.max_row + 1):
        try:
            row = worksheet[row_num]
            
            # استخراج البيانات
            name = str(row[column_map['اسم الشعبة']].value or '').strip()
            grade = str(row[column_map['الصف']].value or '').strip()
            
            if not name or not grade:
                errors.append(f'الصف {row_num}: اسم الشعبة أو الصف مفقود')
                continue
            
            class_type = str(row[column_map['النوع']].value or '').strip()
            stage = str(row[column_map['المرحلة الدراسية']].value or '').strip()
            expected_students = row[column_map['عدد الطلاب المتوقع']].value or 0
            building = str(row[column_map['البناية']].value or '').strip()
            teacher = str(row[column_map['المعلم المسؤول']].value or '').strip()
            
            try:
                expected_students = int(expected_students)
            except:
                expected_students = 0
            
            class_data = {
                'name': name,
                'grade': grade,
                'type': class_type,
                'stage': stage,
                'expected_students': expected_students,
                'building': building,
                'teacher': teacher,
                'is_french': 'فرنسي' in class_type.lower(),
                'created_at': datetime.now().isoformat()
            }
            
            classes_data.append(class_data)
            
        except Exception as e:
            errors.append(f'الصف {row_num}: خطأ في معالجة البيانات - {str(e)}')
    
    success_count = len(classes_data)
    
    response_data = {
        'success': True,
        'count': success_count,
        'type': 'classes',
        'data': classes_data,
        'errors': errors
    }
    
    if errors:
        response_data['warning'] = f'تم رفع {success_count} شعبة بنجاح، مع {len(errors)} خطأ'
    
    return jsonify(response_data)

@excel_bp.route('/download-template/<template_type>', methods=['GET'])
def download_template(template_type):
    """تحميل قوالب Excel"""
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        if template_type == 'students':
            # إنشاء قالب الطلاب
            headers = [
                'الاسم الكامل', 'رقم الهوية', 'تاريخ الميلاد', 'الصف', 'الشعبة',
                'اسم ولي الأمر', 'رقم هاتف ولي الأمر', 'العنوان', 'الجنس', 'البناية'
            ]
            sample_data = [
                ['أحمد محمد علي', '123456789012', '15/03/2007', 'الثالث المتوسط', 'فرنسي', 
                 'محمد علي حسن', '07801234567', 'النجف الأشرف - حي السلام', 'ذكر', 'بنين'],
                ['فاطمة أحمد محمد', '123456789013', '20/05/2007', 'الثالث المتوسط', 'غير فرنسي', 
                 'أحمد محمد حسين', '07801234568', 'النجف الأشرف - حي الجامعة', 'أنثى', 'بنات']
            ]
            filename = 'قالب_الطلاب.xlsx'
            worksheet.title = 'بيانات الطلاب'
            
        elif template_type == 'classes':
            # إنشاء قالب الشعب
            headers = [
                'اسم الشعبة', 'الصف', 'النوع', 'المرحلة الدراسية', 
                'عدد الطلاب المتوقع', 'البناية', 'المعلم المسؤول'
            ]
            sample_data = [
                ['الثالث متوسط أ', 'الثالث المتوسط', 'فرنسي', 'المتوسطة', 25, 'بنين', 'أحمد محمد'],
                ['الثالث متوسط ب', 'الثالث المتوسط', 'غير فرنسي', 'المتوسطة', 30, 'بنات', 'فاطمة علي']
            ]
            filename = 'قالب_الشعب.xlsx'
            worksheet.title = 'بيانات الشعب'
        else:
            return jsonify({'error': 'نوع القالب غير صحيح'}), 400
        
        # إضافة العناوين
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # إضافة البيانات التجريبية
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # تنسيق الأعمدة
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # حفظ في الذاكرة
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'خطأ في إنشاء القالب: {str(e)}'}), 500

